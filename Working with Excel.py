
# coding: utf-8

# # 在Python下使用openpyxl模块读取和修改Excel文件
# Python没有自带openpyxl，需要单独安装。
# 下文代码，是基于编辑本文时openpyxl的最新版本-2.5.9版本。
# 网络上存在的一些基于老版本的代码，在新版本上已无法运行，请予以注意。

# ## 1. 读取Excel文档

# ### 1.1 用openpyxl模块打开Excel文档

# In[1]:


import openpyxl#导入openpyxl库，用于处理excel


# In[2]:


wb = openpyxl.load_workbook('example.xlsx')#打开excel表格
type(wb)#查看表格wb的数据类型


# In[3]:


import os#导入os库，获取/改变工作目录


# In[4]:


os.getcwd()#获取当前工作目录


# ### 1.2 从工作簿中取得工作表

# In[5]:


wb.sheetnames#获取工作簿wb的所有工作表名称


# In[6]:


sheet = wb['Sheet1']#获得工作表
sheet


# In[7]:


type(sheet)#获取工作表Sheet1的数据类型


# In[8]:


sheet.title#获取工作表的名称


# ### 1.3 从工作表中取得单元格

# In[9]:


sheet['A1']#按照名字访问Cell对象


# In[10]:


sheet['A1'].value#获取单元格中保存的值


# In[11]:


b1 = sheet['B1']#将单元格B1赋值给b1


# In[12]:


b1


# In[13]:


b1.row#获取单元格所在行


# In[14]:


b1.column#获取单元格所在列


# In[15]:


b1.coordinate#获取单元格坐标


# In[16]:


sheet.cell(row=1,column=2)#通过数字而非字母获得单元格


# In[17]:


sheet.cell(row=1,column=2).value#获得单元格保存的值


# In[18]:


for i in range(1,8,2):#打印第二列奇数行数据
    print (str(i) + ': ' + sheet.cell(row=i,column=2).value)


# In[19]:


sheet.max_row#获取工作表行数


# In[20]:


sheet.max_column#获取工作表列数


# ### 1.4 列字母和数字之间的转换

# In[21]:


from openpyxl.utils import get_column_letter,column_index_from_string#调用列字母和数字转换函数
get_column_letter(2)#数字转换到字母


# In[22]:


column_index_from_string('AA')#字母转换到数字


# In[23]:


get_column_letter(sheet.max_column)#获取工作表Sheet1最大列数


# ### 1.5 从表中取得行和列

# In[24]:


tuple(sheet['A1':'C3'])#使用tuple()函数，可以看到A1:C3矩形区域中的Cell对象
#这个元组包含3个元组：每个元组代表1行，从指定区域的顶部到底部


# In[25]:


for i in sheet['A1':'C3']:#使用两个for循环，遍历所有单元格的值；外层for循环遍历这个切片的每一行
    for j in i:#内层for循环遍历改行中的每个单元格
        print(j.coordinate,j.value)
    print('---END OF ROW---')


# In[26]:


next(sheet.columns)#获取第一列数据


# In[27]:


sheet['A']#获取第一列数据


# In[28]:


sheet[1]#获取第一行数据


# In[29]:


for i in sheet['B']:#打印第二列数据
    print(i.value)


# ## 2. 写入Excel文档

# ### 2.1 创建并保存Excel文档

# In[30]:


import openpyxl#导入openpyxl模块
wb = openpyxl.Workbook()#创建一个新的工作簿
wb.sheetnames#获取新工作簿的工作表名称


# In[31]:


sheet = wb.active#获得活跃的工作表


# In[32]:


sheet.title#活跃工作表的名称


# In[33]:


sheet.title = 'New Sheet 1'#重命名活跃工作表的名称


# In[34]:


wb.sheetnames#查看活跃工作表名称


# In[35]:


wb.save('test.xlsx')#保存工作簿，如果没有该句，工作簿不会保存


# ### 2.2 创建和删除工作表

# In[36]:


wb.sheetnames#查看所有工作表名称


# In[37]:


wb.create_sheet()#创建工作表，默认名为SheetX，默认是工作簿的最后一个工作表


# In[38]:


wb.sheetnames#查看工作表名称变化


# In[39]:


wb.create_sheet(index=0,title='First Sheet')#通过index、title关键字参数，指定新工作表的索引（位置）和名称


# In[40]:


wb.sheetnames#查看工作表名称变化


# In[41]:


wb.create_sheet(index=2,title='Middle Sheet')#创建索引为2，名称为Middle Sheet的工作表


# In[42]:


wb.sheetnames#查看工作表名称变化


# In[43]:


wb.remove(wb['Middle Sheet'])#删除工作表，remove()方法接受Worksheet对象作为其参数，而不是工作表名称的字符串


# In[44]:


wb.sheetnames#查看工作表名称变化


# In[45]:


wb.save('test1.xlsx')#保存工作表


# ### 2.3 将值写入单元格

# In[46]:


sheet = wb['Sheet']
sheet['A1'] = 'Hello world!'


# In[47]:


sheet['A1'].value


# In[48]:


wb.save('test1.xlsx')#将值保存在原文件中


# ### 2.4 设置单元格的字体风格

# In[49]:


from openpyxl.styles import Font, colors#导入Font()和 colors()函数


# In[50]:


wb['Sheet']['A1'].font = Font(color = colors.RED,italic = True)#设置Sheet工作表A1单元格字体为红色、斜体


# In[51]:


wb.save('test_styled.xlsx')#保存文件


# ### 2.5 Font对象
# Font属性关键字参数
# name：字体名称，诸如'Calibri'
# size：字体大小
# bold：True表示粗体
# italic：True表示斜体

# In[52]:


wb['Sheet']['A2'] = 'Hello world'#给A2单元格赋值
font1 = Font(name='Calibri',size=12,bold=True,italic=False)#将Font对象保存在一个变量中
wb['Sheet']['A2'].font = font1#将A2单元格字体设置为font1形式


# In[53]:


wb.save('test_styled.xlsx')#保存文件


# ### 2.6 公式

# In[54]:


sheet = wb.active#将活跃工作表赋值给sheet
sheet['A1'] = 123
sheet['A2'] = 234
sheet['A3'] = '=SUM(A1:A2)'#在A3单元格书写公式
wb.save('writeFormula.xlsx')#保存文件


# In[55]:


sheet['A3'].value#如果excel有公式，openpyxl默认显示公式而非计算的数值


# In[56]:


wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx',data_only=True)#通过关键字参数data_only可以只显示数值
sheet1 = wbDataOnly.active
sheet1['A3'].value#如果在将公式写入A3单元格后，未打开Excel，则该公式不会计算直至打开Excel并保存。
#所以，公式通过openpyxl写入单元格后，不打开Excel，则该句不会显示计算结果。


# ### 2.7 调整行和列

# #### 2.7.1 设置行高和列宽

# In[57]:


wb = openpyxl.Workbook()#新建工作簿
sheet = wb.active#获取活跃工作表
sheet['A1'] = 'Hello'
sheet['B2'] = 'World'


# In[58]:


sheet.row_dimensions[1].height = 70#设置第一行行高


# In[59]:


sheet.column_dimensions['B'].width = 20#设置第二列列宽


# In[60]:


wb.save('dimensions.xlsx')#保存文件


# #### 2.7.2 合并和拆分单元格

# In[61]:


wb = openpyxl.Workbook()#创建一个工作簿
sheet = wb.active#获取活跃工作表
sheet.merge_cells('A1:D3')#合并A1:D3单元格
sheet['A1'] = '12个单元格合并在一起'#给合并的单元格赋值
sheet.merge_cells('C5:E5')
sheet['C5'] = '3个单元格合并在一起'
wb.save('merged.xlsx')#保存文件


# In[62]:


wb = openpyxl.load_workbook('merged.xlsx')#打开工作簿
sheet = wb.active#获取活跃工作表
sheet.unmerge_cells('A1:D3')#拆分A1:D3单元格
wb.save('merged.xlsx')#保存文件


# #### 2.7.3 冻结窗格

# In[63]:


wb = openpyxl.load_workbook('produceSales.xlsx')#加载工作簿
sheet = wb.active#获取活跃工作表
sheet.freeze_panes = 'A2'#冻结行1
wb.save('freezeExample.xlsx')#保存文件


# #### 2.7.4 图表

# In[64]:


wb = openpyxl.Workbook()#创建工作簿
sheet = wb.active#获取活跃工作表
for i in range(1,21):#创建20个数据
    sheet['A'+str(i)] = i*i

from openpyxl.chart import BarChart,Reference,Series#导入条形图、Reference和Series对象
values = Reference(sheet,min_col=1,min_row=1,max_col=1,max_row=20)#图表的值范围
chart = BarChart()#创建条形图
chart.add_data(values)#在条形图上加载数据
sheet.add_chart(chart,'B2')#在表格上加载条形图
wb.save('sampleChart.xlsx')#保存文件

